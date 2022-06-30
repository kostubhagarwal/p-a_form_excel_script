import os
import pdb
import pandas as pd
from openpyxl import load_workbook

#initiate variables for file and dataframe
cwd = os.getcwd()
files = os.listdir(cwd)
df = pd.DataFrame()

#read function
def readIn (df) :
    df0 = pd.DataFrame()
    df0=df0.append(pd.read_excel(file, sheet_name = SheetN, usecols = "D", skiprows = [0,1,2,3,5,6,8,10,12,14,15,17,19,20,21,25,27,28,29,30,31,32,33,34,35,36,37,38,39]))
    df = pd.concat([df,df0], axis=1) 
    return df

#loop to read through files, and insert as new column to df
for file in files:
    if file.endswith('.xlsm'):
   
        #preparing workbook for openPyXL
        wb = load_workbook(file)
        wb.sheetnames
        sheet = wb["Home"]

        #assessment of which sheet is being used
        x = sheet["R5"].value

        if x == 1:
            SheetN = "TechnicalDeliveryPrincipal"
            df = readIn(df)           

        if x == 2:
            SheetN = "BuisnessDevelopmentPrincipal"
            df = readIn(df)

        if x == 3:
            SheetN = "HybridPrincipal"
            df = readIn(df)

        if x == 4:
            SheetN = "TechnicalDeliveryAssociate"
            df = readIn(df)

        if x == 5:
            SheetN = "BuisnessDevelopmentAssociate"
            df = readIn(df)

        if x == 6:
            SheetN = "HybridAssociate"
            df = readIn(df)
  
#output file
print(df)
df.to_csv('PA_Compiled_Matrix.csv', index=False)


